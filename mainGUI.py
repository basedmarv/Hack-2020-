from tkinter import *
import openpyxl
from PIL import ImageTk, Image

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
def create_list(workbook):
    # wb = openpyxl.load_workbook('pyxl_trial.xlsx')
    ws = workbook.active

    arr = [] 

    for x in range(ws.max_row - 1):
        yuh = ws.cell(x+2,1)
        arr.append(yuh)

    return arr

# def append_list(workbook,val):
#     ws = workbook.active
#     maxR = ws.max_row
#     print(maxR)
#     ws.cell(row=maxR+1, column=1).value = val
#     # ws.append([val,agreement])
    
#     workbook.save(r'D:\Users\Roy Matthew\Documents\Side_Projs\Hack_2020\Hack-2020-\pyxl_trial.xlsx')



def main():
    path = r'D:\Users\Roy Matthew\Documents\Side_Projs\Hack_2020\Hack-2020-\pyxl_trial.xlsx'
    wb = openpyxl.load_workbook(path)
    # ws = wb.active

    yessir = create_list(wb)
    # append_list(wb,"YUH")

    top = Tk()

    # window description
    top.title('The Newlyweds')
    top.geometry("500x400")

    # c = Canvas(top, height=600, width=500)
    # img = ImageTk.PhotoImage(Image.open("bg.jpg"))
    tobeWeds = Label(top, fg="red", text="Mr. and Mrs. Wedding", font="Times 16 italic").pack()
    
    # text_try = Text(top, height = 2, width = 50)
    # text_try.tag_configure("left", justify='left')
    # text_try.pack()
    # text_try.insert(END, yessir[0].value)
    # text_try = Text(top, fg = "red", text=yessir[0], font="Times 11", wraplength=250).pack()

    # listbox = Listbox(top, fg="red", font="Times 11")
    # listbox.pack(fill=BOTH, expand=1)

    maxR = wb.active.max_row
    # print(maxR)
    count = 10


    if maxR < 10:
        for x in range(len(yessir)):
            text_try = Text(top, height = 2, width = 50)
            text_try.tag_configure("left", justify='left')
            text_try.pack()
            text_try.insert(END, yessir[x].value)
    else: 
        while count > 0:
            text_try = Text(top, height = 2, width = 75)
            text_try.tag_configure("left", justify='left')
            text_try.pack()
            text_try.insert(END, yessir[maxR-count-1].value)
            count = count - 1

        # listbox.insert(END, yessir[x].value)


    top.mainloop() 
 
 
if __name__ == "__main__":
    main()