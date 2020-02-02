import Credentials
from twilio import twiml
from twilio.rest import Client
from twilio.twiml.messaging_response import MessagingResponse
from flask import Flask, request, redirect
import openpyxl
from openpyxl import Workbook


account_sid = Credentials.get_sid()
auth_token = Credentials.get_auth()
path = Credentials.path

app = Flask(__name__)

def update_xl(message_body):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    lastRow = sheet.max_row
    sheet.cell(column = 1, row = lastRow + 1, value="{0}".format(message_body))

    wb.save(filename = path)



@app.route("/sms", methods=['GET','POST'])
def sms_reply():

    #Retrieve actual message
    message_body = request.form['Body']

    #Update Excel Sheet
    update_xl(message_body)

    resp = MessagingResponse()
    resp.message("Hi!! Thank you for the message! I'm sure they'll enjoy it")

    return str(resp)

if __name__ == "__main__":
    app.run(debug=True)